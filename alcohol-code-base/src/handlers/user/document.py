import asyncio
import logging
from datetime import datetime
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import Optional

from aiogram import Router, F
from aiogram.types import Message, Document, PhotoSize
from aiogram.fsm.context import FSMContext
from aiogram.filters import Command

from config import get_settings
from enums.sheets import Sheet
from keyboards.default.document import DocumentKeyboards
from services import SheetsMappingService
from states.document import DocumentProcessing

from services.n8n_service import N8nService
from services.gemini_service import GeminiService
from services.google_service import GoogleService
from services.perplexity_service import PerplexityService

from models.processing_result import ProcessingResult
from keyboards.default.basic import BasicButtons
from utils.item_extended import extend_items

# Настройка логирования
logger = logging.getLogger(__name__)

settings = get_settings()

# Инициализация сервисов
N8N_WEBHOOK_URL = settings.n8n.webhook_url
GEMINI_API_KEY = settings.gemini.api_key
GOOGLE_CREDENTIALS_PATH = settings.google.credentials_path
DELEGATION_EMAIL = settings.google.delegated_user_email
PERPLEXITY_API_KEY = settings.perplexity.api_key

n8n_service = N8nService(N8N_WEBHOOK_URL) if N8N_WEBHOOK_URL else None
gemini_service = GeminiService(GEMINI_API_KEY) if GEMINI_API_KEY else None
perplexity_service = PerplexityService(PERPLEXITY_API_KEY) if PERPLEXITY_API_KEY else None

google_service = GoogleService(GOOGLE_CREDENTIALS_PATH, DELEGATION_EMAIL)

document_router = Router()

sheets_mapping_service = SheetsMappingService()


@document_router.message(Command("analyze"))
async def start_document_analysis(message: Message, state: FSMContext) -> None:
    """
    Начинает процесс анализа документа с выбора проекта.

    Args:
        message: Сообщение пользователя
        state: Состояние FSM
    """
    try:
        # Получаем список доступных проектов
        all_mappings = sheets_mapping_service.get_all_mappings()

        if not all_mappings:
            await message.answer(
                "❌ <b>Нет доступных проектов</b>\n\n"
                "Обратитесь к администратору для настройки проектов.",
                reply_markup=BasicButtons.back(),
            )
            return

        project_names = list(all_mappings.keys())

        await state.set_state(DocumentProcessing.waiting_for_project_selection)

        await message.answer(
            "🏗️ <b>Выбор проекта</b>\n\n"
            "Выберите проект, к которому относится документ для анализа:",
            reply_markup=DocumentKeyboards.project_selection(project_names),
        )

    except Exception as e:
        logger.error(f"Ошибка при запуске анализа документа: {e}")
        await message.answer(
            "❌ Произошла ошибка при запуске анализа. Попробуйте позже.",
            reply_markup=BasicButtons.back(),
        )


@document_router.message(DocumentProcessing.waiting_for_project_selection)
async def process_project_selection(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает выбор проекта пользователем.

    Args:
        message: Сообщение с выбранным проектом
        state: Состояние FSM
    """
    try:
        selected_project = message.text.strip()

        # Проверяем, нажата ли кнопка "Отмена"
        if selected_project == "🚫 Отмена":
            await state.clear()
            await message.answer(
                "🚫 <b>Обработка отменена</b>\n\n"
                "Вы можете начать заново нажав 'Анализ документа' или отправив команду /analyze",
                reply_markup=BasicButtons.main_menu(),
            )
            return

        # Проверяем, что выбранный проект существует
        all_mappings = sheets_mapping_service.get_all_mappings()

        if selected_project not in all_mappings:
            await message.answer(
                "❌ Неверный выбор проекта. Пожалуйста, выберите из предложенных вариантов:",
                reply_markup=DocumentKeyboards.project_selection(
                    list(all_mappings.keys())
                ),
            )
            return

        # Сохраняем выбранный проект
        await state.update_data(selected_project=selected_project)
        await state.set_state(DocumentProcessing.waiting_for_document_type_selection)

        await message.answer(
            f"✅ <b>Выбран проект:</b> {selected_project}\n\n"
            "📋 <b>Тип документа</b>\n\n"
            "Выберите, что описывает предоставляемый документ:",
            reply_markup=DocumentKeyboards.document_type_selection(),
        )

    except Exception as e:
        logger.error(f"Ошибка при обработке выбора проекта: {e}")
        await message.answer(
            "❌ Произошла ошибка. Попробуйте выбрать проект снова.",
            reply_markup=BasicButtons.back(),
        )
        await state.clear()


@document_router.message(DocumentProcessing.waiting_for_document_type_selection)
async def process_document_type_selection(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает выбор типа документа пользователем.

    Args:
        message: Сообщение с выбранным типом документа
        state: Состояние FSM
    """
    try:
        selected_text = message.text.strip()

        # Проверяем, нажата ли кнопка "Отмена"
        if selected_text == "🚫 Отмена":
            await state.clear()
            await message.answer(
                "🚫 <b>Обработка отменена</b>\n\n"
                "Вы можете начать заново нажав 'Анализ документа' или отправив команду /analyze",
                reply_markup=BasicButtons.main_menu(),
            )
            return

        # Определяем тип документа на основе выбора
        if selected_text == "📦 Товары":
            document_type = Sheet.MATERIALS
            type_description = "товары"
        elif selected_text == "🔧 Услуги":
            document_type = Sheet.JOBS
            type_description = "услуги"
        else:
            await message.answer(
                "❌ Неверный выбор типа документа. Пожалуйста, выберите из предложенных вариантов:",
                reply_markup=DocumentKeyboards.document_type_selection(),
            )
            return

        # Сохраняем выбранный тип документа
        await state.update_data(document_type=document_type)
        await state.set_state(DocumentProcessing.waiting_for_document)

        # Получаем данные о выбранном проекте
        state_data = await state.get_data()
        selected_project = state_data.get("selected_project", "Неизвестный проект")

        await message.answer(
            f"✅ <b>Настройки анализа:</b>\n"
            f"🏗️ Проект: {selected_project}\n"
            f"📋 Тип: {type_description}\n\n"
            "📄 <b>Отправьте документ или фото для анализа</b>\n\n"
            "После обработки вы получите"
            " ссылку на итоговый документ",
            reply_markup=BasicButtons.cancel(),
        )

    except Exception as e:
        logger.error(f"Ошибка при обработке выбора типа документа: {e}")
        await message.answer(
            "❌ Произошла ошибка. Попробуйте выбрать тип документа снова.",
            reply_markup=BasicButtons.back(),
        )
        await state.clear()


@document_router.message(DocumentProcessing.waiting_for_document, F.photo)
async def process_document_photo(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает загруженное изображение документа.

    Args:
        message: Сообщение с изображением
        state: Состояние FSM
    """
    try:
        await state.set_state(DocumentProcessing.processing)

        processing_msg = await message.answer(
            "⏳ <b>Анализирую изображение ...</b>\n\n"
            "Это может занять несколько минут.\n"
            "Пожалуйста, подождите.",
            reply_markup=BasicButtons.cancel(),
        )

        # Получаем изображение наилучшего качества
        photo: PhotoSize = message.photo[-1]

        # Скачиваем изображение
        file_info = await message.bot.get_file(photo.file_id)
        file_extension = ".jpg"  # По умолчанию Telegram отдает JPG
        file_name = f"image_{message.from_user.id}_{photo.file_id}{file_extension}"
        file_path = Path(f"temp_{file_name}")

        await message.bot.download_file(file_info.file_path, file_path)

        # Получаем данные из состояния
        state_data = await state.get_data()
        selected_project = state_data.get("selected_project")
        document_type = state_data.get("document_type")

        # Обрабатываем изображение через Gemini
        result = await _process_document_pipeline(
            file_path,
            file_name,
            message.from_user.id,
            message.from_user.first_name,
            message.from_user.username,
            selected_project=selected_project,
            document_type=document_type,
            is_image=True,
        )

        # Удаляем временный файл
        try:
            file_path.unlink()
        except Exception as e:
            logger.warning(f"Не удалось удалить временный файл {file_path}: {e}")

        # Удаляем сообщение о обработке
        try:
            await processing_msg.delete()
        except Exception:
            pass

        # Отправляем результат
        if result.success:
            await _send_success_result(message, result, state)
        else:
            await _send_error_result(message, result.error_message, state)

    except Exception as e:
        logger.error(f"Ошибка при обработке изображения: {e}")
        await message.answer(
            "❌ Произошла ошибка при обработке изображения. Попробуйте позже.",
            reply_markup=BasicButtons.back(),
        )
        await state.clear()


@document_router.message(DocumentProcessing.waiting_for_document, F.document)
async def process_document_file(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает загруженный файл документа.

    Args:
        message: Сообщение с документом
        state: Состояние FSM
    """
    document: Document = message.document

    try:
        await state.set_state(DocumentProcessing.processing)

        # Определяем тип файла по MIME-type или расширению
        is_image = _is_image_file(document)

        processing_msg = await message.answer(
            "⏳ <b>Обрабатываю документ...</b>\n\n"
            "Это может занять несколько минут.\n"
            "Пожалуйста, подождите.",
            reply_markup=BasicButtons.cancel(),
        )

        # Скачиваем файл
        file_info = await message.bot.get_file(document.file_id)
        file_path = Path(f"temp_{message.from_user.id}_{document.file_name}")

        await message.bot.download_file(file_info.file_path, file_path)

        # Получаем данные из состояния
        state_data = await state.get_data()
        selected_project = state_data.get("selected_project")
        document_type = state_data.get("document_type")

        # Обрабатываем документ через n8n или Gemini
        result = await _process_document_pipeline(
            file_path,
            document.file_name,
            message.from_user.id,
            message.from_user.first_name,
            message.from_user.username,
            selected_project=selected_project,
            document_type=document_type,
            is_image=is_image,
        )

        # Удаляем временный файл
        try:
            file_path.unlink()
        except Exception as e:
            logger.warning(f"Не удалось удалить временный файл {file_path}: {e}")

        # Удаляем сообщение о обработке
        try:
            await processing_msg.delete()
        except Exception:
            pass

        # Отправляем результат
        if result.success:
            await _send_success_result(message, result, state)
        else:
            await _send_error_result(message, result.error_message, state)

    except Exception as e:
        logger.error(f"Ошибка при обработке документа: {e}")
        await message.answer(
            "❌ Произошла ошибка при обработке документа. Попробуйте позже.",
            reply_markup=BasicButtons.back(),
        )
        await state.clear()


@document_router.message(DocumentProcessing.waiting_for_document)
async def invalid_document_format(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает неправильный формат сообщения.

    Args:
        message: Сообщение пользователя
        state: Состояние FSM
    """
    await message.answer(
        "❌ Пожалуйста, отправьте документ файлом или изображением.\n\n"
        "📄 <b>Поддерживаемые форматы:</b>\n"
        "• Изображения: JPG, JPEG, PNG, GIF, BMP, WEBP\n\n"
        "• Документы: PDF, DOCX\n"
        "💡 <b>Способы отправки:</b>\n"
        "• Как фото (сжатое) - для быстрой обработки\n"
        "• Как файл (без сжатия) - для лучшего качества",
        reply_markup=BasicButtons.cancel(),
    )


@document_router.message(F.text == "◀️Назад")
async def handle_back_in_document_flow(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает кнопку "Назад" в процессе работы с документами.

    Args:
        message: Сообщение пользователя
        state: Состояние FSM
    """
    try:
        current_state = await state.get_state()

        if (
            current_state
            == DocumentProcessing.waiting_for_document_type_selection.state
        ):
            # Возвращаемся к выбору проекта
            all_mappings = sheets_mapping_service.get_all_mappings()
            project_names = list(all_mappings.keys())

            await state.set_state(DocumentProcessing.waiting_for_project_selection)
            await message.answer(
                "🏗️ <b>Выбор проекта</b>\n\n"
                "Выберите проект, к которому относится документ для анализа:",
                reply_markup=DocumentKeyboards.project_selection(project_names),
            )

        elif current_state == DocumentProcessing.waiting_for_document.state:
            # Возвращаемся к выбору типа документа
            await state.set_state(
                DocumentProcessing.waiting_for_document_type_selection
            )

            state_data = await state.get_data()
            selected_project = state_data.get("selected_project", "Неизвестный проект")

            await message.answer(
                f"✅ <b>Выбран проект:</b> {selected_project}\n\n"
                "📋 <b>Тип документа</b>\n\n"
                "Выберите, что описывает предоставляемый документ:",
                reply_markup=DocumentKeyboards.document_type_selection(),
            )

        else:
            # Для всех остальных состояний возвращаемся в главное меню
            await state.clear()
            from handlers.user.start import command_start_handler

            await command_start_handler(message)

    except Exception as e:
        logger.error(f"Ошибка при обработке кнопки 'Назад' в документном потоке: {e}")
        await message.answer(
            "❌ Произошла ошибка при навигации.",
            reply_markup=BasicButtons.main_menu(),
        )
        await state.clear()


@document_router.message(DocumentProcessing.processing)
async def handle_processing_state_messages(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает сообщения во время обработки документа.

    Args:
        message: Сообщение пользователя
        state: Состояние FSM
    """
    try:
        if message.text == "🚫 Отмена":
            await cancel_processing(message, state)
        else:
            await message.answer(
                "⏳ <b>Документ обрабатывается...</b>\n\n"
                "Пожалуйста, подождите завершения обработки.\n"
                "Для отмены нажмите кнопку 'Отмена'.",
                reply_markup=BasicButtons.cancel(),
            )

    except Exception as e:
        logger.error(f"Ошибка при обработке сообщения в состоянии processing: {e}")


def _is_image_file(document: Document) -> bool:
    """
    Определяет, является ли документ изображением.

    Args:
        document: Объект документа Telegram

    Returns:
        True, если файл является изображением
    """
    try:
        # Проверяем MIME-type
        if document.mime_type and document.mime_type.startswith("image/"):
            return True

        # Проверяем расширение файла
        if document.file_name:
            file_extension = Path(document.file_name).suffix.lower()
            image_extensions = {
                ".jpg",
                ".jpeg",
                ".png",
                ".gif",
                ".bmp",
                ".webp",
                ".tiff",
                ".tif",
            }
            return file_extension in image_extensions

        return False

    except Exception as e:
        logger.warning(f"Ошибка при определении типа файла: {e}")
        return False


async def _process_document_pipeline(
    file_path: Path,
    file_name: str,
    user_id: int,
    first_name: str,
    nik_name: str = None,
    selected_project: str = None,
    document_type: Sheet = None,
    is_image: bool = False,
) -> ProcessingResult:
    """
    Выполняет полный цикл обработки документа с fallback на локальное сохранение.

    Args:
        file_path: Путь к файлу
        file_name: Имя файла
        user_id: ID пользователя
        first_name: Имя пользователя
        nik_name: Никнейм пользователя
        selected_project: Выбранный проект
        document_type: Тип документа (товары или услуги)
        is_image: Является ли файл изображением

    Returns:
        Результат обработки
    """
    try:
        # Шаг 1: Анализ документа
        logger.info(
            f"Начинаем анализ {'изображения' if is_image else 'документа'} {file_name} "
            f"для пользователя {first_name} ({user_id}), проект: {selected_project}, "
            f"тип: {document_type.value if document_type else 'не_указан'}"
        )

        if not gemini_service:
            return ProcessingResult(
                success=False,
                error_message="Сервис анализа документов (Gemini) недоступен",
            )

        if is_image:
            # Анализ изображения через Gemini
            analysis = await gemini_service.analyze_invoice_image(
                file_path,
                document_type,
            )
        else:
            # Анализ документов, не картинок, через Gemini
            analysis = await gemini_service.analyze_document(file_path, document_type)

        if not analysis:
            return ProcessingResult(
                success=False,
                error_message=f"Не удалось проанализировать {'изображение' if is_image else 'документ'}",
            )

        # Шаг 1.5: Обогащение данных рыночной информацией через Perplexity
        if perplexity_service and analysis.items:
            try:
                logger.info(
                    f"Начинаем обогащение {len(analysis.items)} позиций "
                    f"рыночными данными для пользователя {user_id}"
                )

                extended_items = await extend_items(
                    items=analysis.items,
                    document_type=document_type,
                    perplexity_service=perplexity_service,
                    max_concurrent=2,
                )

                # Заменяем оригинальные items на обогащенные
                analysis.items = extended_items

                logger.info(
                    f"Успешно обогащено {len(extended_items)} позиций "
                    f"рыночными данными для пользователя {user_id}"
                )

            except Exception as e:
                logger.warning(
                    f"Не удалось обогатить данные рыночной информацией: {e}. "
                    f"Продолжаем с оригинальными данными."
                )
                # Продолжаем работу с оригинальными данными
        else:
            if not perplexity_service:
                logger.info(
                    "Perplexity сервис недоступен, пропускаем обогащение данных"
                )
            elif not analysis.items:
                logger.info(
                    "Нет позиций для обогащения в документе"
                )

        return await _process_with_google_services(
            analysis,
            user_id,
            first_name,
            file_path,
            file_name,
            selected_project,
            document_type,
        )

    except Exception as e:
        logger.error(f"Ошибка в pipeline обработки документа: {e}")
        return ProcessingResult(
            success=False, error_message=f"Внутренняя ошибка: {str(e)}"
        )


async def _check_google_services_availability() -> bool:
    """
    Проверяет доступность Google сервисов.

    Returns:
        True, если Google сервисы доступны
    """
    try:
        # Проверяем возможность создания клиентов Google
        # drive_service = await google_service._get_drive_service()
        # sheets_client = await google_service._get_sheets_client()
        #
        # return drive_service is not None and sheets_client is not None

        return True

    except Exception as e:
        logger.warning(f"Google сервисы недоступны: {e}")
        return False


async def _process_with_google_services(
    analysis,
    user_id: int,
    first_name: str,
    file_path: Path,
    file_name: str,
    selected_project: str = None,
    document_type: Sheet = None,
) -> ProcessingResult:
    """
    Обрабатывает документ с использованием Google сервисов.

    Args:
        analysis: Результат анализа документа
        user_id: ID пользователя
        file_path: Путь к файлу
        file_name: Имя файла

    Returns:
        Результат обработки
    """
    try:
        # Получаем ID таблицы для выбранного проекта
        spreadsheet_id = None
        if selected_project:
            spreadsheet_id = sheets_mapping_service.get_sheet_id(selected_project)
            if not spreadsheet_id:
                logger.warning(f"Не найден ID таблицы для проекта '{selected_project}'")

        # Шаг 2: Добавление данных в Google таблицу
        logger.info(
            f"Добавляем данные в Google таблицу для пользователя {user_id}, "
            f"проект: {selected_project}, тип: {document_type.value if document_type else 'не указан'}"
        )
        sheet_url = await google_service.add_data_to_spreadsheet(
            analysis,
            user_id,
            first_name,
            spreadsheet_id=spreadsheet_id,
            sheet_name=document_type.value,
        )

        if not sheet_url:
            # Если не удалось добавить данные, переключаемся на локальное сохранение
            logger.warning(
                "Не удалось добавить данные в Google таблицу. "
                "Переключаемся на локальное сохранение."
            )
            return await _process_with_local_storage(
                analysis,
                user_id,
                first_name,
                file_path,
                file_name,
                selected_project,
                document_type,
            )

        return ProcessingResult(
            success=True,
            analysis=analysis,
            google_sheet_url=sheet_url,
            selected_project=selected_project,
            document_type=document_type,
        )

    except Exception as e:
        logger.error(f"Ошибка при обработке через Google сервисы: {e}")
        # Fallback на локальное сохранение
        logger.info("Переключаемся на локальное сохранение")
        return await _process_with_local_storage(
            analysis,
            user_id,
            first_name,
            file_path,
            file_name,
            selected_project,
            document_type,
        )


async def _process_with_local_storage(
    analysis,
    user_id: int,
    first_name: str,
    file_path: Path,
    file_name: str,
    selected_project: str = None,
    document_type: Sheet = None,
) -> ProcessingResult:
    """
    Обрабатывает документ с локальным сохранением.

    Args:
        analysis: Результат анализа документа
        user_id: ID пользователя
        file_path: Путь к файлу
        file_name: Имя файла

    Returns:
        Результат обработки
    """
    try:
        # Шаг 2: Создание Excel таблицы
        logger.info(
            f"Создаем Excel таблицу локально для пользователя {user_id}, "
            f"проект: {selected_project}, тип: {document_type.value if document_type else 'не указан'}"
        )
        excel_path = await _create_local_excel(
            analysis,
            user_id,
            first_name,
            selected_project,
            document_type,
        )

        if not excel_path:
            return ProcessingResult(
                success=False, error_message="Не удалось создать Excel таблицу"
            )

        # Шаг 3: Сохранение JSON в локальную папку json_data
        logger.info(f"Сохраняем JSON локально для пользователя {user_id}")
        json_path = await _save_local_json(
            analysis,
            user_id,
            selected_project,
            document_type,
        )

        if not json_path:
            return ProcessingResult(
                success=False, error_message="Не удалось сохранить JSON файл"
            )

        # Шаг 4: Подготовка Excel-таблицы для передачи пользователю
        logger.info(f"Подготавливаем Excel файл для отправки пользователю {user_id}")

        return ProcessingResult(
            success=True,
            analysis=analysis,
            local_excel_path=excel_path,
            local_json_path=json_path,
            is_local_processing=True,
            selected_project=selected_project,
            document_type=document_type,
        )

    except Exception as e:
        logger.error(f"Ошибка при локальной обработке: {e}")
        return ProcessingResult(
            success=False, error_message=f"Ошибка локального сохранения: {str(e)}"
        )


async def _create_local_excel(
    analysis,
    user_id: int,
    first_name: str,
    selected_project: str = None,
    document_type: Sheet = None,
) -> Optional[Path]:
    """
    Создает Excel таблицу с результатами анализа локально.

    Args:
        analysis: Результат анализа документа
        user_id: ID пользователя

    Returns:
        Путь к созданному Excel файлу или None в случае ошибки
    """
    try:
        # Создаем директорию для Excel файлов, если её нет
        excel_dir = Path("excel_data")
        excel_dir.mkdir(exist_ok=True)

        # Создаем имя файла
        project_prefix = f"{selected_project}_" if selected_project else ""
        type_suffix = f"_{document_type.value}" if document_type else ""
        excel_filename = (
            f"{project_prefix}analysis_{analysis.document_number or 'document'}"
            f"{type_suffix}_{user_id}.xlsx"
        )
        excel_path = excel_dir / excel_filename

        # Создаем новую книгу Excel
        workbook = Workbook()
        worksheet = workbook.active

        # Устанавливаем название листа в зависимости от типа документа
        if document_type:
            worksheet.title = document_type.value
        else:
            worksheet.title = "неправильно"

        # Настраиваем стили
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        # Заполняем заголовки в зависимости от типа документа
        if document_type == Sheet.JOBS:
            headers = [
                "№",
                "Дата",
                "Найменування послуги",
                "Од вим",
                "К-сть",
                "Цiна",
                "Сума",
                "Виконавець",
                "Примiтки",
                "Дата заповнення",
                "TelegramID / First Name",
            ]
        else:  # MATERIALS или по умолчанию
            headers = [
                "№",
                "Дата",
                "Найменування",
                "Од вим",
                "К-сть",
                "Цiна",
                "Сума",
                "Постачальник",
                "Примiтки",
                "Дата заповнення",
                "TelegramID / First Name",
            ]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment

        # Заполняем данные
        row = 2
        for item in analysis.items:
            worksheet.cell(row=row, column=1, value=analysis.document_number or "")
            worksheet.cell(row=row, column=2, value=analysis.date)
            worksheet.cell(row=row, column=3, value=item.name)
            worksheet.cell(row=row, column=4, value=item.unit.lower())
            worksheet.cell(row=row, column=5, value=item.quantity)
            worksheet.cell(row=row, column=6, value=item.price)
            worksheet.cell(row=row, column=7, value=item.total)
            worksheet.cell(row=row, column=8, value=analysis.supplier)
            worksheet.cell(row=row, column=9, value="")
            worksheet.cell(
                row=row, column=10, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            worksheet.cell(row=row, column=11, value=f"{str(user_id)} / {first_name}")
            row += 1

        # Добавляем итоговую строку
        if analysis.total_amount:
            total_row = row
            worksheet.cell(row=total_row, column=4, value="ИТОГО:").font = header_font
            worksheet.cell(
                row=total_row, column=7, value=analysis.total_amount
            ).font = header_font

        # Автоподбор ширины колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        workbook.save(excel_path)
        logger.info(f"Excel файл успешно создан: {excel_path}")

        return excel_path

    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}")
        return None


async def _save_local_json(
    analysis,
    user_id: int,
    selected_project: str = None,
    document_type: Sheet = None,
) -> Optional[Path]:
    """
    Сохраняет результат анализа в JSON файл локально.

    Args:
        analysis: Результат анализа документа
        user_id: ID пользователя
        selected_project: Выбранный проект
        document_type: Тип документа

    Returns:
        Путь к созданному JSON файлу или None в случае ошибки
    """
    try:
        # Создаем директорию для JSON файлов, если её нет
        json_dir = Path("json_data")
        json_dir.mkdir(exist_ok=True)

        # Создаем имя файла с учетом проекта и типа
        project_prefix = f"{selected_project}_" if selected_project else ""
        type_suffix = f"_{document_type.value}" if document_type else ""
        json_filename = (
            f"{project_prefix}analysis_{analysis.document_number or 'document'}"
            f"{type_suffix}_{user_id}.json"
        )
        json_path = json_dir / json_filename

        # Подготавливаем данные для сохранения с дополнительной информацией
        json_data = analysis.model_dump()
        json_data.update(
            {
                "processing_info": {
                    "user_id": user_id,
                    "selected_project": selected_project,
                    "document_type": document_type.value if document_type else None,
                    "processed_at": datetime.now().isoformat(),
                    "processing_method": "local",
                }
            }
        )

        # Сохраняем данные в JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON файл успешно создан: {json_path}")
        return json_path

    except Exception as e:
        logger.error(f"Ошибка при сохранении JSON файла: {e}")
        return None


async def _send_success_result(
    message: Message, result: ProcessingResult, state: FSMContext
) -> None:
    """
    Отправляет успешный результат обработки документа.

    Args:
        message: Сообщение пользователя
        result: Результат обработки
        state: Состояние FSM
    """
    try:
        analysis = result.analysis

        # Формируем краткую информацию о документе
        doc_info = []
        if analysis.document_type:
            doc_info.append(f"📋 Тип: {analysis.document_type}")
        if analysis.document_number:
            doc_info.append(f"📄 Номер: {analysis.document_number}")
        if analysis.date:
            doc_info.append(f"📅 Дата: {analysis.date}")
        if analysis.supplier:
            doc_info.append(f"🏢 Поставщик: {analysis.supplier}")
        if analysis.customer:
            doc_info.append(f"🏪 Покупатель: {analysis.customer}")

        doc_summary = (
            "\n".join(doc_info) if doc_info else "Основная информация не распознана"
        )

        # Формируем информацию о товарах
        items_count = len(analysis.items)
        items_info = f"📦 Товаров: {items_count} шт."

        if analysis.total_amount:
            items_info += f"\n💰 Общая сумма: {analysis.total_amount} грн."

        if result.is_local_processing:
            # Локальная обработка - отправляем файлы напрямую
            response_text = (
                "✅ <b>Документ успешно обработан!</b>\n"
                "⚠️ <i>Google сервисы недоступны, файлы сохранены локально</i>\n\n"
                f"{doc_summary}\n\n"
                f"{items_info}\n\n"
                "📊 <b>Результаты:</b>\n"
                "• Excel таблица (отправляется файлом)\n"
            )

            await message.answer(
                response_text,
                reply_markup=BasicButtons.back(),
            )

            # Отправляем Excel файл
            if result.local_excel_path and result.local_excel_path.exists():
                try:
                    from aiogram.types import FSInputFile

                    excel_file = FSInputFile(
                        result.local_excel_path,
                        filename=f"analysis_{analysis.document_number or 'document'}.xlsx",
                    )
                    await message.answer_document(
                        excel_file, caption="📊 Excel таблица с результатами анализа"
                    )

                    # Удаляем временный файл после отправки
                    try:
                        result.local_excel_path.unlink()
                    except Exception as e:
                        logger.warning(f"Не удалось удалить временный Excel файл: {e}")

                except Exception as e:
                    logger.error(f"Ошибка при отправке Excel файла: {e}")
                    await message.answer("❌ Не удалось отправить Excel файл")

            # Отправляем JSON файл
            # if result.local_json_path and result.local_json_path.exists():
            #     try:
            #         from aiogram.types import FSInputFile
            #
            #         json_file = FSInputFile(
            #             result.local_json_path,
            #             filename=f"analysis_{analysis.document_number or 'document'}.json",
            #         )
            #         await message.answer_document(
            #             json_file, caption="📄 JSON данные с результатами анализа"
            #         )
            #
            #         # Удаляем временный файл после отправки
            #         try:
            #             result.local_json_path.unlink()
            #         except Exception as e:
            #             logger.warning(f"Не удалось удалить временный JSON файл: {e}")
            #
            #     except Exception as e:
            #         logger.error(f"Ошибка при отправке JSON файла: {e}")
            #         await message.answer("❌ Не удалось отправить JSON файл")

        else:
            # Google обработка - отправляем ссылки
            response_text = (
                "✅ <b>Документ успешно обработан!</b>\n\n"
                f"{doc_summary}\n\n"
                f"{items_info}\n\n"
                "📊 <b>Данные добавлены в общую таблицу:</b>\n"
                f"• <a href='{result.google_sheet_url}'>Посмотреть таблицу</a>"
            )

            await message.answer(
                response_text,
                reply_markup=BasicButtons.back(),
                disable_web_page_preview=True,
            )

        # Отправляем детальную информацию о товарах, если их немного
        if analysis.items:
            items_text = "📋 <b>Детали по товарам:</b>\n\n"

            for i, item in enumerate(analysis.items, 1):
                items_text += (
                    f"{i}. <b>{item.name}</b>\n"
                    f"   Количество: {item.quantity} {item.unit}\n"
                    f"   Цена: {item.price} грн.\n"
                    f"   Сумма: {item.total} грн.\n\n"
                )

        await message.answer(items_text)

        await state.clear()
        logger.info(
            f"Успешно отправлен результат {'локальной' if result.is_local_processing else 'Google'} "
            f"обработки для пользователя {message.from_user.id}"
        )

    except Exception as e:
        logger.error(f"Ошибка при отправке успешного результата: {e}")
        await message.answer(
            "✅ Документ обработан, но произошла ошибка при отображении результатов.",
            reply_markup=BasicButtons.back(),
        )
        await state.clear()


async def _send_error_result(
    message: Message, error_message: Optional[str], state: FSMContext
) -> None:
    """
    Отправляет сообщение об ошибке.

    Args:
        message: Сообщение пользователя
        error_message: Сообщение об ошибке
        state: Состояние FSM
    """
    try:
        response_text = (
            "❌ <b>Ошибка при обработке документа</b>\n\n"
            f"Причина: {error_message or 'Неизвестная ошибка'}\n\n"
            "Попробуйте:\n"
            "• Загрузить документ лучшего качества\n"
            "• Проверить формат файла\n"
            "• Повторить попытку позже"
        )

        await message.answer(response_text, reply_markup=BasicButtons.back())
        await state.clear()

    except Exception as e:
        logger.error(f"Ошибка при отправке сообщения об ошибке: {e}")


@document_router.message(F.text == "🚫 Отмена")
async def cancel_processing(message: Message, state: FSMContext) -> None:
    """
    Отменяет текущую операцию обработки документа.

    Args:
        message: Сообщение пользователя
        state: Состояние FSM
    """
    try:
        await state.clear()
        await message.answer(
            "🚫 <b>Обработка отменена</b>\n\n"
            "Вы можете начать заново, отправив команду /analyze",
            reply_markup=BasicButtons.main_menu(),
        )

    except Exception as e:
        logger.error(f"Ошибка при отмене обработки: {e}")
        await message.answer(
            "Обработка отменена",
            reply_markup=BasicButtons.main_menu(),
        )
